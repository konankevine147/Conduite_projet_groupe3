import time
import re
import logging
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException,
    StaleElementReferenceException, WebDriverException,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────────────────────

BASE_URL = (
    "https://www.welcometothejungle.com/fr/jobs"
    "?refinementList%5Boffices.country_code%5D%5B%5D=FR"
    "&query=data"

# Les données sont stockées dans un fichier excel

)
EXCEL_PATH = "wttj_jobs_data.xlsx"
HEADLESS   = False   # WTTJ bloque le headless → fenêtre Chrome visible
DELAY      = 2.0
MAX_OFFERS = 30

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("scraper.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)


# ── Modèle ─────────────────────────────────────────────────────────────────────

@dataclass
class JobOffer:
    title:       str
    company:     str
    city:        str
    salary:      Optional[str] = None
    description: Optional[str] = None
    profile:     Optional[str] = None
    url:         Optional[str] = None

# ── Excel ──────────────────────────────────────────────────────────────────────

COLUMNS    = ["Intitulé du poste", "Entreprise", "Ville", "Salaire", "Description", "Profil recherché", "URL"]
COL_WIDTHS = [35, 25, 20, 20, 60, 60, 50]

HEADER_FILL   = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ROW_FILL_ODD  = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
ROW_FILL_EVEN = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
CELL_FONT     = Font(name="Arial", size=10)
BORDER_SIDE   = Side(style="thin", color="BDD7EE")
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)


def init_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Offres emplois"
    ws.freeze_panes = "A2"
    for col_idx, (header, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    return wb, ws


def append_offer(ws, job: JobOffer, row: int):
    values = [job.title, job.company, job.city, job.salary or "", job.description or "", job.profile or "", job.url or ""]
    fill = ROW_FILL_ODD if row % 2 != 0 else ROW_FILL_EVEN
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font      = CELL_FONT
        cell.fill      = fill
        cell.border    = CELL_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 80


# ── Navigateur ─────────────────────────────────────────────────────────────────

def build_driver():
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--start-maximized")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"},
    )
    return driver


# ── Helpers ────────────────────────────────────────────────────────────────────

def try_selectors(root, selectors: list) -> str:
    for sel in selectors:
        try:
            el = root.find_element(By.CSS_SELECTOR, sel)
            txt = el.text.strip()
            if txt:
                return txt
        except NoSuchElementException:
            pass
    return ""


def clean_text(text: str) -> str:
    """Supprime les blocs parasites ajoutés par WTTJ."""
    if not text:
        return ""
    text = re.sub(r"D['']autres offres.*", "", text, flags=re.DOTALL)
    text = re.sub(r"Voir plus\s*$", "", text, flags=re.MULTILINE)
    return text.strip()


def extract_company_from_url(url: str) -> str:
    m = re.search(r"/companies/([^/]+)/", url)
    return m.group(1).replace("-", " ").title() if m else ""


def extract_city_from_url(url: str) -> str:
    m = re.search(r"_([a-zà-ü\-]+)(?:_[A-Z]{2,}|$)", url)
    return m.group(1).replace("-", " ").title() if m else ""


# ── Collecte des liens ─────────────────────────────────────────────────────────

def get_job_links(driver) -> list:
    log.info("🌐 Chargement de la page de recherche...")
    driver.get(BASE_URL)
    try:
        WebDriverWait(driver, 25).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/jobs/']"))
        )
    except TimeoutException:
        log.error("⏱️ Timeout — page non chargée.")
        with open("debug_list.html", "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        return []
    time.sleep(3)

    links = []
    for a in driver.find_elements(By.CSS_SELECTOR, "a[href]"):
        try:
            href = a.get_attribute("href") or ""
            if re.search(r"/companies/.+/jobs/", href) and href not in links:
                links.append(href)
                if len(links) >= MAX_OFFERS:
                    break
        except StaleElementReferenceException:
            continue

    log.info("🔗 %d liens collectés", len(links))
    return links


# ── Scraping d'une offre ───────────────────────────────────────────────────────

def extract_sections(driver) -> tuple:
    desc, prof = [], []
    DESC_KW = {"poste", "mission", "description", "rôle", "role", "responsabilit"}
    PROF_KW = {"profil", "vous", "candidat", "compétence", "expérience", "requis"}

    try:
        for section in driver.find_elements(By.CSS_SELECTOR, "section, [data-testid*='section']"):
            try:
                header = section.find_element(By.CSS_SELECTOR, "h2,h3,h4").text.lower()
            except NoSuchElementException:
                continue
            content = section.text.strip()
            if any(k in header for k in DESC_KW):
                desc.append(content)
            elif any(k in header for k in PROF_KW):
                prof.append(content)
    except Exception:
        pass

    if not desc:
        fallback = try_selectors(driver, [
            "[data-testid*='description']", "article", "main"
        ])
        if fallback:
            desc.append(fallback)

    return clean_text("\n\n".join(desc)), clean_text("\n\n".join(prof))


def scrape_offer(driver, url: str) -> Optional[JobOffer]:
    try:
        driver.get(url)
    except WebDriverException as e:
        log.warning("❌ Chargement impossible : %s", e)
        return None

    try:
        WebDriverWait(driver, 20).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "h1, h2, main"))
        )
    except TimeoutException:
        log.warning("⏱️ Timeout sur : %s", url)
        return None

    time.sleep(2)

    title    = try_selectors(driver, ["h1", "h2", "[data-testid*='title']"])
    company  = extract_company_from_url(url)
    city     = extract_city_from_url(url)

    salary = try_selectors(driver, [
        "[data-testid*='salary']", "[data-testid*='remuneration']",
        "[data-testid*='compensation']", "[aria-label*='alaire']",
    ])
    if not salary:
        try:
            body_text = driver.find_element(By.TAG_NAME, "body").text
            m = re.search(r"\d[\d\s]*[kK€][\s]*[-–à]?\s*\d*[\d\s]*[kK€]?", body_text)
            if m:
                salary = m.group(0).strip()
        except Exception:
            pass

    description, profile = extract_sections(driver)

    if not title or not company:
        log.warning("⚠️  title='%s' company='%s' → ignorée : %s", title, company, url)
        return None

    log.info("   ✔ %s | %s | %s", title[:50], company[:30], city)
    return JobOffer(
        title=title, company=company, city=city,
        salary=salary or None, description=description or None,
        profile=profile or None, url=url,
    )


# ── Main ───────────────────────────────────────────────────────────────────────

def run():
    wb, ws = init_workbook()
    driver = build_driver()
    total  = 0

    log.info("🚀 Démarrage — Chrome va s'ouvrir, ne pas le fermer !")

    try:
        links = get_job_links(driver)
        if not links:
            log.error("❌ Aucun lien trouvé.")
            return

        for url in links:
            try:
                job = scrape_offer(driver, url)
                if job:
                    total += 1
                    append_offer(ws, job, row=total + 1)
                    wb.save(EXCEL_PATH)
                    log.info("✅ [%d/%d] %s — %s | %s | Salaire: %s",
                             total, MAX_OFFERS, job.title, job.company,
                             job.city, job.salary or "N/A")
                time.sleep(DELAY)
            except KeyboardInterrupt:
                raise
            except Exception as e:
                log.error("❌ Erreur : %s", e)

    except KeyboardInterrupt:
        log.info("⛔ Interruption manuelle.")
    finally:
        driver.quit()
        wb.save(EXCEL_PATH)
        log.info("🎉 Terminé — %d offres sauvegardées dans %s", total, EXCEL_PATH)


if __name__ == "__main__":
    run()
